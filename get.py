import json, os, random, copy, threading, multiprocessing, re
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter
from rules import apply_rules, sort_classes
from paths import MSC_FILE, BACKEND_FILE
DAYS            = ["Mon","Tue","Wed","Thu","Fri","Sat"]
PERIODS_PER_DAY = 8
DOWNLOADS       = os.path.join(os.path.expanduser("~"), "Downloads")


BLOCKED_FILL = {"PET","LIBRARY","ART","DL","VE",
                "MPT","CCA","CS PRACTICAL","PHY/CHEM PRACTICAL",
                "BIO PRACTICAL","MATHS/CS/HINDI"}

# Subjects allowed in rescue fill — only real academic subjects
# teacher must already be assigned to that class in msc.json
RESCUE_PRIORITY_SUBJECTS = {
    "PHYSICS", "CHEMISTRY", "BIOLOGY", "ECONOMICS",
    "COMMERCE/BUSINESS STD.", "ENGLISH", "MATHS", "SCIENCE",
    "SST", "HINDI", "SANSKRIT", "CS", "AI"
}

ENGLISH_ONCE_CLASSES = {"11a","11b","11c","12a","12b","12c"}

# -------------------------------------------------------
# MATHS CONFIG
# 11A is lead for the 11-group, 12A is lead for the 12-group.
# Each group gets MATHS_PERIODS_TOTAL (9) blocks total.
# The two groups MUST NOT share the same (day, period) slot.
# Max MATHS_MAX_PER_DAY maths blocks per day per group.
# -------------------------------------------------------
MATHS_PAIRED = {
    "11a": ["11b","11c"],
    "12a": ["12b","12c"]
}
MATHS_BUSY_TEACHERS = ["JAYA","KIRAN","SOJU"]
MATHS_LEAD_TEACHER  = MATHS_BUSY_TEACHERS[0]
MATHS_PERIODS_TOTAL = 9          # ← changed from 8 to 9
MATHS_MAX_PER_DAY   = 2

PHY_CHEM_TEACHERS = {
    "11A": ["RAJANI","SAJIMON"],
    "12B": ["RAJANI","SAJIMON"],
    "11B": ["LIJI MATHEW","POOJA SIHAG"],
    "12A": ["LIJI MATHEW","POOJA SIHAG"],
}
CS_PRACTICAL_TEACHER  = "SOJU"
BIO_PRACTICAL_TEACHER = "BINDU C"

PHASE2_MAX_PER_DAY = 3
MAX_ATTEMPTS       = 50

SMART_FILL_ITERS   = 5
SWAP_OPT_ITERS     = 50
PIPELINE_ROUNDS    = 2


# -------------------------------------------------------
# TEXT CLEANER
# -------------------------------------------------------

def sanitize(text):
    if text is None:
        return ""
    text = str(text)
    if text.strip() in {"—", "-", "--"}:
        return ""
    result = []
    for ch in text:
        code = ord(ch)
        if 32 <= code <= 126:
            result.append(ch)
        elif code in (0x2014, 0x2013): result.append("-")
        elif code in (0x2018, 0x2019): result.append("'")
        elif code in (0x201C, 0x201D): result.append('"')
        elif code == 0x2026:           result.append("...")
        elif code == 0x00A0:           result.append(" ")
    return "".join(result)


def norm(cls):
    return str(cls).strip().lower()


# -------------------------------------------------------
# TEACHER → CLASS ASSIGNMENT CACHE
# -------------------------------------------------------

def build_teacher_class_map(msc_data):
    tcmap = {}
    for teacher, info in msc_data.items():
        if not teacher or teacher.strip() in {"—", "-", ""}:
            continue
        assigned = {norm(c) for c in info.get("classes", {}).keys()}
        tcmap[teacher] = assigned
    return tcmap


def teacher_teaches_class(teacher, cls, teacher_class_map):
    return norm(cls) in teacher_class_map.get(teacher, set())


# -------------------------------------------------------
# SUBJECT COUNT HELPERS
# -------------------------------------------------------

def subject_count_day(table, day, subject):
    return sum(
        1 for p in range(PERIODS_PER_DAY)
        if table[day][p] and table[day][p]["subject"] == subject
    )


def english_allowed(cls, table, day):
    count = subject_count_day(table, day, "ENGLISH")
    if norm(cls) in {norm(c) for c in ENGLISH_ONCE_CLASSES}:
        return count == 0
    return count < 2


def subject_placement_allowed(cls, table, day, subject, preferred_max=1):
    count = subject_count_day(table, day, subject)
    if count >= 2:
        return False
    if count >= preferred_max:
        return False
    if subject == "ENGLISH":
        return english_allowed(cls, table, day)
    return True


def subject_placement_allowed_phase2(cls, table, day, subject):
    count = subject_count_day(table, day, subject)
    if count >= PHASE2_MAX_PER_DAY:
        return False
    if subject == "ENGLISH":
        return english_allowed(cls, table, day)
    return True


# -------------------------------------------------------
# GET VALID FILL SUBJECTS FOR A CLASS
# -------------------------------------------------------

def get_fill_subjects(cls, msc_data, teacher_class_map):
    allowed = []
    for teacher, info in msc_data.items():
        if not teacher or teacher.strip() in {"—", "-", ""}:
            continue
        subject = sanitize(info.get("subject", ""))
        if not subject:
            continue
        if subject.upper() in {b.upper() for b in BLOCKED_FILL}:
            continue
        if not teacher_teaches_class(teacher, cls, teacher_class_map):
            continue
        allowed.append((subject, teacher))
    return allowed


# -------------------------------------------------------
# BACKEND HELPER
# -------------------------------------------------------

def mark_teacher_busy(teacher_name, day, period, cls_name,
                      teacher_avail_global, backend_grid):
    if teacher_name in teacher_avail_global:
        teacher_avail_global[teacher_name][day][period] = False
    if teacher_name in backend_grid:
        backend_grid[teacher_name][day][period] = cls_name


# -------------------------------------------------------
# SYNC PRACTICAL AVAILABILITY
# -------------------------------------------------------

def sync_practical_availability(timetable, teacher_avail_global, backend_grid):
    for cls, grid in timetable.items():
        cls_upper = cls.upper()
        for day in range(len(DAYS)):
            for period in range(PERIODS_PER_DAY):
                cell = grid[day][period]
                if not cell:
                    continue
                subject = cell.get("subject", "").upper()

                if subject == "CS PRACTICAL":
                    mark_teacher_busy(CS_PRACTICAL_TEACHER, day, period, cls,
                                      teacher_avail_global, backend_grid)
                elif subject == "PHY/CHEM PRACTICAL":
                    for t in PHY_CHEM_TEACHERS.get(cls_upper, []):
                        mark_teacher_busy(t, day, period, cls,
                                          teacher_avail_global, backend_grid)
                elif subject == "BIO PRACTICAL":
                    mark_teacher_busy(BIO_PRACTICAL_TEACHER, day, period, cls,
                                      teacher_avail_global, backend_grid)
                else:
                    teacher = cell.get("teacher", "")
                    if teacher and teacher.strip() not in {"—", "-", "", "MATHS/CS/HINDI"}:
                        for t in [x.strip() for x in teacher.replace(",", "&").split("&")]:
                            mark_teacher_busy(t, day, period, cls,
                                              teacher_avail_global, backend_grid)


# -------------------------------------------------------
# MATHS BLOCK PLACER
#
# Key changes vs original:
#   1. MATHS_PERIODS_TOTAL = 9 (was 8)
#   2. The two groups (11-group, 12-group) track their OWN
#      used slots in `group_used_slots[group_key]`.
#      A slot is forbidden for a group only if THAT GROUP
#      already uses it — the two groups are now fully
#      independent of each other's slot choices.
#   3. All MATHS_BUSY_TEACHERS are still marked busy for
#      every placed slot (they attend both groups' periods).
#   4. Equal distribution: we try MATHS_MAX_PER_DAY=2 per
#      day; with 9 slots over 6 days that gives
#      [2,2,2,1,1,1] or [2,2,2,2,1,0] etc., which is the
#      best possible spread. The spread_key sort ensures
#      days with fewer blocks are preferred first.
# -------------------------------------------------------

def place_maths_blocks(timetable, teacher_avail_global, backend_grid):
    cls_map = {norm(c): c for c in timetable.keys()}

    # Track, per group, which (day,period) slots have been used
    group_used_slots = {}   # group_key -> set of (day,period)

    # All slots ever used by ANY maths group — teachers can't double-book
    all_teacher_busy_slots = set()

    for lead_raw, paired_raw in MATHS_PAIRED.items():
        group_key = lead_raw          # e.g. "11a" or "12a"
        lead      = cls_map.get(norm(lead_raw))
        paired    = [cls_map[norm(p)] for p in paired_raw if norm(p) in cls_map]

        if lead is None:
            continue

        group_used_slots[group_key] = set()
        placed_slots = []
        day_counts   = {d: 0 for d in range(len(DAYS))}

        candidates = [(d, p) for d in range(len(DAYS)) for p in range(PERIODS_PER_DAY)]
        random.shuffle(candidates)

        def spread_key(dp):
            return day_counts[dp[0]]

        def try_place_maths(day, period):
            # Already placed for this group?
            if (day, period) in group_used_slots[group_key]:
                return False
            # Teachers busy this slot (any group)?
            if (day, period) in all_teacher_busy_slots:
                return False
            # Day cap for this group
            if day_counts[day] >= MATHS_MAX_PER_DAY:
                return False
            # Timetable cells must be free
            if timetable[lead][day][period] is not None:
                return False
            if any(timetable[p_cls][day][period] is not None for p_cls in paired):
                return False
            # Teacher availability (all three must be free)
            if any(
                t in teacher_avail_global and not teacher_avail_global[t][day][period]
                for t in MATHS_BUSY_TEACHERS
            ):
                return False

            # ── Place the block ──
            timetable[lead][day][period] = {
                "subject": "MATHS",
                "teacher": MATHS_LEAD_TEACHER
            }
            for p_cls in paired:
                timetable[p_cls][day][period] = {
                    "subject": "MATHS/CS/HINDI",
                    "teacher": ", ".join(MATHS_BUSY_TEACHERS)
                }

            all_maths_classes = [lead] + paired
            all_classes_str   = ", ".join(sorted(all_maths_classes))
            for t in MATHS_BUSY_TEACHERS:
                if t in teacher_avail_global:
                    teacher_avail_global[t][day][period] = False
                if t in backend_grid:
                    backend_grid[t][day][period] = all_classes_str

            placed_slots.append((day, period))
            group_used_slots[group_key].add((day, period))
            all_teacher_busy_slots.add((day, period))
            day_counts[day] += 1
            return True

        # ── Primary pass: prefer days with fewer blocks (spread evenly) ──
        attempts = 0
        while len(placed_slots) < MATHS_PERIODS_TOTAL and attempts < 3000:
            attempts += 1
            candidates.sort(key=spread_key)
            for day, period in candidates:
                if try_place_maths(day, period):
                    break

        # ── Fallback: relax spread preference, just fill remaining ──
        if len(placed_slots) < MATHS_PERIODS_TOTAL:
            for day, period in candidates:
                if len(placed_slots) >= MATHS_PERIODS_TOTAL:
                    break
                try_place_maths(day, period)


# -------------------------------------------------------
# PLACE ONE TASK
# -------------------------------------------------------

def try_place_task(cls, teacher, subject, timetable,
                   teacher_avail_global, backend_grid, preferred_max):
    days    = list(range(len(DAYS)))
    periods = list(range(PERIODS_PER_DAY))
    random.shuffle(days)
    random.shuffle(periods)

    for day in days:
        if not subject_placement_allowed(cls, timetable[cls], day, subject, preferred_max):
            continue
        for period in periods:
            if teacher not in teacher_avail_global:
                return False
            if (timetable[cls][day][period] is None
                    and teacher_avail_global[teacher][day][period]):
                timetable[cls][day][period] = {"subject": subject, "teacher": teacher}
                mark_teacher_busy(teacher, day, period, cls,
                                  teacher_avail_global, backend_grid)
                return True
    return False


# -------------------------------------------------------
# COUNT HOW MANY TIMES A TASK IS ALREADY PLACED
# -------------------------------------------------------

def placed_count(cls, teacher, subject, timetable):
    return sum(
        1 for d in range(len(DAYS))
        for p in range(PERIODS_PER_DAY)
        if timetable[cls][d][p]
        and timetable[cls][d][p]["subject"] == subject
        and timetable[cls][d][p]["teacher"] == teacher
    )


# -------------------------------------------------------
# SMART GAP FILL
# -------------------------------------------------------

def smart_fill(timetable, msc_data, teacher_avail_global, backend_grid,
               classes, teacher_class_map):
    for preferred_max in [1, 2]:
        for _iteration in range(SMART_FILL_ITERS):
            changed = False
            for cls in classes:
                table        = timetable[cls]
                fill_options = get_fill_subjects(cls, msc_data, teacher_class_map)
                random.shuffle(fill_options)

                for day in range(len(DAYS)):
                    for period in range(PERIODS_PER_DAY):
                        if table[day][period] is not None:
                            continue
                        for subject, teacher in fill_options:
                            if teacher not in teacher_avail_global:
                                continue
                            if not teacher_avail_global[teacher][day][period]:
                                continue
                            if not subject_placement_allowed(cls, table, day, subject, preferred_max):
                                continue
                            table[day][period] = {"subject": subject, "teacher": teacher}
                            mark_teacher_busy(teacher, day, period, cls,
                                              teacher_avail_global, backend_grid)
                            changed = True
                            break
            if not changed:
                break


# -------------------------------------------------------
# SMART GAP FILL — PHASE 2
# -------------------------------------------------------

def smart_fill_phase2(timetable, msc_data, teacher_avail_global, backend_grid,
                      classes, teacher_class_map):
    for _iteration in range(SMART_FILL_ITERS):
        changed = False
        for cls in classes:
            table        = timetable[cls]
            fill_options = get_fill_subjects(cls, msc_data, teacher_class_map)
            random.shuffle(fill_options)

            for day in range(len(DAYS)):
                for period in range(PERIODS_PER_DAY):
                    if table[day][period] is not None:
                        continue
                    for subject, teacher in fill_options:
                        if teacher not in teacher_avail_global:
                            continue
                        if not teacher_avail_global[teacher][day][period]:
                            continue
                        if not subject_placement_allowed_phase2(cls, table, day, subject):
                            continue
                        table[day][period] = {"subject": subject, "teacher": teacher}
                        mark_teacher_busy(teacher, day, period, cls,
                                          teacher_avail_global, backend_grid)
                        changed = True
                        break
        if not changed:
            break


# -------------------------------------------------------
# TARGETED SWAP
# -------------------------------------------------------

def swap_optimizer(timetable, msc_data, teacher_avail_global, backend_grid,
                   classes, teacher_class_map):
    for _iteration in range(SWAP_OPT_ITERS):
        pending_swaps = []

        for cls in classes:
            table          = timetable[cls]
            valid_subjects = {s for s, _ in get_fill_subjects(cls, msc_data, teacher_class_map)}

            for day in range(len(DAYS)):
                for period in range(PERIODS_PER_DAY):
                    if table[day][period] is not None:
                        continue
                    for other in classes:
                        if other == cls:
                            continue
                        cell = timetable[other][day][period]
                        if not cell:
                            continue
                        subject = cell["subject"]
                        teacher = cell["teacher"]
                        if subject.upper() in {b.upper() for b in BLOCKED_FILL}:
                            continue
                        if subject not in valid_subjects:
                            continue
                        if not teacher_teaches_class(teacher, cls, teacher_class_map):
                            continue
                        if not subject_placement_allowed(cls, table, day, subject, 2):
                            continue
                        pending_swaps.append((cls, other, day, period, cell))
                        break

        if not pending_swaps:
            break

        for cls, other, day, period, cell in pending_swaps:
            if timetable[cls][day][period] is not None:
                continue
            if timetable[other][day][period] is None:
                continue
            teacher = cell["teacher"]
            timetable[cls][day][period]   = cell
            timetable[other][day][period] = None
            if teacher in backend_grid:
                backend_grid[teacher][day][period] = cls


# -------------------------------------------------------
# DEEP SWAP
# -------------------------------------------------------

def deep_swap(timetable, msc_data, teacher_avail_global, backend_grid,
              classes, teacher_class_map):
    for cls in classes:
        table          = timetable[cls]
        valid_subjects = {s for s, _ in get_fill_subjects(cls, msc_data, teacher_class_map)}
        moved          = False

        for day in range(len(DAYS)):
            if moved:
                break
            for period in range(PERIODS_PER_DAY):
                if moved:
                    break
                if table[day][period] is not None:
                    continue
                for src_day in range(len(DAYS)):
                    if moved:
                        break
                    for src_period in range(PERIODS_PER_DAY):
                        src_cell = table[src_day][src_period]
                        if not src_cell:
                            continue
                        teacher = src_cell["teacher"]
                        subject = src_cell["subject"]
                        if subject.upper() in {b.upper() for b in BLOCKED_FILL}:
                            continue
                        if subject not in valid_subjects:
                            continue
                        if not teacher_teaches_class(teacher, cls, teacher_class_map):
                            continue
                        if teacher not in teacher_avail_global:
                            continue
                        if not teacher_avail_global[teacher][day][period]:
                            continue
                        if not subject_placement_allowed(cls, table, day, subject, 2):
                            continue

                        table[day][period]          = src_cell
                        table[src_day][src_period]  = None
                        teacher_avail_global[teacher][day][period]         = False
                        teacher_avail_global[teacher][src_day][src_period] = True
                        if teacher in backend_grid:
                            backend_grid[teacher][day][period]         = cls
                            backend_grid[teacher][src_day][src_period] = None
                        moved = True
                        break


# -------------------------------------------------------
# FORCED FILL
# -------------------------------------------------------

def forced_fill(timetable, msc_data, teacher_avail_global, backend_grid,
                classes, teacher_class_map):
    for cls in classes:
        table        = timetable[cls]
        fill_options = get_fill_subjects(cls, msc_data, teacher_class_map)

        for day in range(len(DAYS)):
            for period in range(PERIODS_PER_DAY):
                if table[day][period] is not None:
                    continue
                random.shuffle(fill_options)
                for subject, teacher in fill_options:
                    if teacher not in teacher_avail_global:
                        continue
                    if not teacher_avail_global[teacher][day][period]:
                        continue
                    if subject_count_day(table, day, subject) >= 2:
                        continue
                    if subject == "ENGLISH" and not english_allowed(cls, table, day):
                        continue
                    table[day][period] = {"subject": subject, "teacher": teacher}
                    mark_teacher_busy(teacher, day, period, cls,
                                      teacher_avail_global, backend_grid)
                    break


# -------------------------------------------------------
# RESCUE FILL
# Absolute last resort. For every remaining empty slot:
#   - Try every teacher assigned to that class in msc.json
#   - Only requirement: teacher is FREE at that slot
#   - No per-day subject cap, no English restriction
#   - Only skip: practicals, CCA, MPT, DL, VE, ART, MATHS/CS/HINDI
#   - PET and LIBRARY are allowed
# -------------------------------------------------------

RESCUE_SKIP = {"ART", "DL", "VE", "MPT", "CCA",
               "CS PRACTICAL", "PHY/CHEM PRACTICAL", "BIO PRACTICAL",
               "MATHS/CS/HINDI"}

RESCUE_PRIORITY = [
    "PHYSICS", "CHEMISTRY", "BIOLOGY", "ECONOMICS",
    "COMMERCE/BUSINESS STD.", "CS", "MATHS", "SCIENCE",
    "SST", "HINDI", "SANSKRIT", "AI", "ENGLISH", "PET", "LIBRARY",
]

def rescue_fill(timetable, msc_data, teacher_avail_global, backend_grid,
                classes, teacher_class_map):

    for cls in classes:
        table = timetable[cls]

        # Every teacher assigned to this class, skip only hard-blocked subjects
        candidates = []
        for teacher, info in msc_data.items():
            if not teacher or teacher.strip() in {"—", "-", ""}:
                continue
            subject = sanitize(info.get("subject", ""))
            if not subject:
                continue
            if subject.upper() in {s.upper() for s in RESCUE_SKIP}:
                continue
            if not teacher_teaches_class(teacher, cls, teacher_class_map):
                continue
            candidates.append((subject, teacher))

        def pri(st):
            s = st[0].upper()
            for i, p in enumerate(RESCUE_PRIORITY):
                if s == p:
                    return i
            return len(RESCUE_PRIORITY)

        candidates.sort(key=pri)

        for day in range(len(DAYS)):
            for period in range(PERIODS_PER_DAY):
                if table[day][period] is not None:
                    continue
                for subject, teacher in candidates:
                    # Only check: is teacher free?
                    if teacher not in teacher_avail_global:
                        continue
                    if not teacher_avail_global[teacher][day][period]:
                        continue
                    table[day][period] = {"subject": subject, "teacher": teacher}
                    mark_teacher_busy(teacher, day, period, cls,
                                      teacher_avail_global, backend_grid)
                    break

def build_empty_report(timetable, classes):
    lines = []
    for cls in classes:
        for d, day in enumerate(DAYS):
            for p in range(PERIODS_PER_DAY):
                if timetable[cls][d][p] is None:
                    lines.append(f"  Class {cls}  |  {day}  |  Period {p+1}")
    if not lines:
        return ""
    return "⚠ Unfilled slots:\n" + "\n".join(lines)


# -------------------------------------------------------
# CORE COMPUTATION
# -------------------------------------------------------

def _run_attempts(msc_data, backend_data_existing, classes,
                  teacher_class_map, progress_queue, max_attempts):
    try:
        best_timetable = None
        best_empty     = 999999
        best_backend   = None

        for attempt in range(max_attempts):

            timetable = {cls: [[None]*PERIODS_PER_DAY for _ in DAYS] for cls in classes}

            teacher_avail_global = {
                teacher: [[True]*PERIODS_PER_DAY for _ in DAYS]
                for teacher in msc_data.keys()
                if teacher and teacher.strip() not in {"—", "-", ""}
            }

            backend_grid = {
                teacher: [[None]*PERIODS_PER_DAY for _ in DAYS]
                for teacher in teacher_avail_global
            }

            timetable = apply_rules(timetable, msc_data, backend_data_existing)
            sync_practical_availability(timetable, teacher_avail_global, backend_grid)

            # ── Pre-place MATHS_BUSY_TEACHERS' non-paired classes FIRST ──
            # e.g. JAYA teaches 8C maths (7 periods) — ALL must be placed
            # before place_maths_blocks() runs so those slots are already
            # marked busy and maths blocks can't land on the same slot.
            pre_place_classes = []  # (teacher, cls, subject, total_count)
            for teacher in MATHS_BUSY_TEACHERS:
                info = msc_data.get(teacher, {})
                subject = sanitize(info.get("subject", ""))
                if not subject:
                    continue
                for cls, count in info.get("classes", {}).items():
                    if norm(cls) in {norm(k) for k in MATHS_PAIRED}:
                        continue
                    paired_values = [norm(v) for vlist in MATHS_PAIRED.values() for v in vlist]
                    if norm(cls) in paired_values:
                        continue
                    pre_place_classes.append((teacher, cls, subject, count))

            # Place ALL periods for each pre-task class before maths blocks
            for teacher, cls, subject, count in pre_place_classes:
                for preferred_max in [1, 2]:
                    while placed_count(cls, teacher, subject, timetable) < count:
                        placed_before = placed_count(cls, teacher, subject, timetable)
                        try_place_task(cls, teacher, subject,
                                       timetable, teacher_avail_global, backend_grid,
                                       preferred_max=preferred_max)
                        if placed_count(cls, teacher, subject, timetable) == placed_before:
                            break  # no progress, move to next preferred_max
            # ─────────────────────────────────────────────────────────────

            place_maths_blocks(timetable, teacher_avail_global, backend_grid)

            pre_place_set = {(t, c) for t, c, s, n in pre_place_classes}

            tasks = []
            for teacher, info in msc_data.items():
                if not teacher or teacher.strip() in {"—", "-", ""}:
                    continue
                subject = sanitize(info.get("subject", ""))
                if not subject:
                    continue
                for cls, count in info.get("classes", {}).items():
                    if subject.upper() == "MATHS" and norm(cls) in {norm(k) for k in MATHS_PAIRED}:
                        continue
                    if subject.upper() == "MATHS/CS/HINDI":
                        continue
                    # Skip already pre-placed
                    if (teacher, cls) in pre_place_set:
                        continue
                    for _ in range(count):
                        tasks.append({"teacher": teacher, "class": cls, "subject": subject})

            random.shuffle(tasks)

            def task_priority(t):
                if (t["subject"].upper() == "ENGLISH"
                        and norm(t["class"]) in {norm(c) for c in ENGLISH_ONCE_CLASSES}):
                    return 0
                return 1

            tasks.sort(key=task_priority)

            for preferred_max in [1, 2]:
                for task in tasks:
                    cls     = task["class"]
                    teacher = task["teacher"]
                    subject = task["subject"]
                    needed  = msc_data.get(teacher, {}).get("classes", {}).get(cls, 0)
                    if placed_count(cls, teacher, subject, timetable) >= needed:
                        continue
                    try_place_task(cls, teacher, subject, timetable,
                                   teacher_avail_global, backend_grid, preferred_max)

            for _ in range(PIPELINE_ROUNDS):
                smart_fill(timetable, msc_data, teacher_avail_global, backend_grid,
                           classes, teacher_class_map)
                swap_optimizer(timetable, msc_data, teacher_avail_global, backend_grid,
                               classes, teacher_class_map)
                deep_swap(timetable, msc_data, teacher_avail_global, backend_grid,
                          classes, teacher_class_map)

            forced_fill(timetable, msc_data, teacher_avail_global, backend_grid,
                        classes, teacher_class_map)
            smart_fill(timetable, msc_data, teacher_avail_global, backend_grid,
                       classes, teacher_class_map)

            empty_after_phase1 = sum(
                1 for cls in classes
                for d in range(len(DAYS))
                for p in range(PERIODS_PER_DAY)
                if timetable[cls][d][p] is None
            )

            if empty_after_phase1 > 0:
                smart_fill_phase2(timetable, msc_data, teacher_avail_global, backend_grid,
                                  classes, teacher_class_map)
                swap_optimizer(timetable, msc_data, teacher_avail_global, backend_grid,
                               classes, teacher_class_map)

            # ── RESCUE FILL ──────────────────────────────────────
            # Runs after all normal passes. Fills any remaining
            # empty slot using the class's own assigned teachers,
            # no per-day subject cap. Teacher must still be free.
            rescue_fill(timetable, msc_data, teacher_avail_global, backend_grid,
                        classes, teacher_class_map)
            # ─────────────────────────────────────────────────────

            empty_count = sum(
                1 for cls in classes
                for d in range(len(DAYS))
                for p in range(PERIODS_PER_DAY)
                if timetable[cls][d][p] is None
            )

            if empty_count < best_empty:
                best_empty     = empty_count
                best_timetable = copy.deepcopy(timetable)
                best_backend   = copy.deepcopy(backend_grid)

            progress_queue.put({
                "attempt": attempt + 1,
                "total":   max_attempts,
                "empty":   best_empty,
            })

            if best_empty == 0:
                break

        progress_queue.put({
            "done":      True,
            "timetable": best_timetable,
            "backend":   best_backend,
            "empty":     best_empty,
        })

    except Exception as exc:
        progress_queue.put({"error": str(exc)})


# -------------------------------------------------------
# EXCEL STYLES HELPER
# -------------------------------------------------------

def _make_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_fill():
    return PatternFill("solid", fgColor="2E4057")   # dark navy

def _alt_row_fill(i):
    return PatternFill("solid", fgColor="F0F4F8" if i % 2 == 0 else "FFFFFF")

def _practical_fill():
    return PatternFill("solid", fgColor="FFF3CD")   # soft amber

def _maths_fill():
    return PatternFill("solid", fgColor="D4EDDA")   # soft green

def _cell_fill():
    return PatternFill("solid", fgColor="E8F4FD")   # soft blue

def _style_header_row(ws, row_num, num_cols):
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border    = _make_border()
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font      = hdr_font
        cell.fill      = _header_fill()
        cell.alignment = hdr_align
        cell.border    = border

def _style_data_cell(ws, row_num, col_num, subject="", row_i=0):
    cell   = ws.cell(row=row_num, column=col_num)
    border = _make_border()
    cell.border    = border
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    subj_upper = subject.upper()
    if "PRACTICAL" in subj_upper:
        cell.fill = _practical_fill()
        cell.font = Font(size=9, color="856404")
    elif subj_upper in {"MATHS", "MATHS/CS/HINDI"}:
        cell.fill = _maths_fill()
        cell.font = Font(size=9, color="155724")
    elif subj_upper in {"MPT", "CCA"}:
        cell.fill = PatternFill("solid", fgColor="F8D7DA")
        cell.font = Font(size=9, color="721C24")
    elif subject:
        cell.fill = _cell_fill()
        cell.font = Font(size=9, color="1A4A6B")
    else:
        cell.fill = _alt_row_fill(row_i)
        cell.font = Font(size=9, color="888888")


# -------------------------------------------------------
# EXCEL BUILDER
# -------------------------------------------------------

def _build_excel(timetable, backend_grid, best_empty, msc_data, classes):

    # ── 1. Build backend teacher map ────────────────────────────────────────────
    backend_out = {}

    def maths_lead_teachers(cls):
        return [MATHS_LEAD_TEACHER]

    def maths_paired_teachers(cls):
        return [t for t in MATHS_BUSY_TEACHERS if t != MATHS_LEAD_TEACHER]

    practical_teachers = {
        "CS PRACTICAL":       lambda cls: [CS_PRACTICAL_TEACHER],
        "BIO PRACTICAL":      lambda cls: [BIO_PRACTICAL_TEACHER],
        "PHY/CHEM PRACTICAL": lambda cls: PHY_CHEM_TEACHERS.get(cls.upper(), []),
        "MATHS":              maths_lead_teachers,
        "MATHS/CS/HINDI":     maths_paired_teachers,
    }

    MATHS_BLOCK_SUBJECTS = {"MATHS", "MATHS/CS/HINDI"}

    teacher_slot_map = {
        t: [[None]*PERIODS_PER_DAY for _ in DAYS]
        for t in msc_data.keys()
        if t and t.strip() not in {"—", "-", ""}
    }

    def write_slot(teacher, day, period, display):
        if teacher not in teacher_slot_map:
            return
        existing = teacher_slot_map[teacher][day][period]
        if existing is None:
            teacher_slot_map[teacher][day][period] = display
        elif display not in existing.split(", "):
            teacher_slot_map[teacher][day][period] = existing + ", " + display

    for cls in timetable:
        m = re.match(r'^(\d+)[A-Za-z]$', cls)
        grade_label = m.group(1) if m else cls
        for day in range(len(DAYS)):
            for period in range(PERIODS_PER_DAY):
                cell = timetable[cls][day][period]
                if not cell:
                    continue
                subject      = cell.get("subject", "").upper()
                cell_teacher = cell.get("teacher", "")
                if subject in practical_teachers:
                    owners  = practical_teachers[subject](cls)
                    display = grade_label if subject in MATHS_BLOCK_SUBJECTS else cls
                    for t in owners:
                        if subject in MATHS_BLOCK_SUBJECTS:
                            # Overwrite with grade label only — never append
                            if t in teacher_slot_map:
                                teacher_slot_map[t][day][period] = display
                        else:
                            write_slot(t, day, period, display)
                else:
                    owners = [t.strip() for t in cell_teacher.replace(",", "&").split("&") if t.strip()]
                    for t in owners:
                        # If this is a maths-busy teacher and the slot already
                        # holds a grade label (digits only), don't overwrite it
                        if (t in MATHS_BUSY_TEACHERS
                                and t in teacher_slot_map
                                and teacher_slot_map[t][day][period] is not None
                                and str(teacher_slot_map[t][day][period]).isdigit()):
                            continue
                        write_slot(t, day, period, cls)

    for teacher, info in msc_data.items():
        if not teacher or teacher.strip() in {"—", "-", ""}:
            continue
        backend_out[teacher] = {
            "subject": sanitize(info.get("subject", "")),
            "grid":    teacher_slot_map.get(teacher, [[None]*PERIODS_PER_DAY for _ in DAYS])
        }

    with open(BACKEND_FILE, "w") as f:
        json.dump(backend_out, f, indent=4)

    # ── 2. Build Excel workbook ─────────────────────────────────────────────────
    DOWNLOADS   = os.path.join(os.path.expanduser("~"), "Downloads")
    xlsx_path   = os.path.join(DOWNLOADS, "timetable.xlsx")
    wb = Workbook()
    first_sheet = wb.active   # keep the default sheet, rename it later or replace it

    period_labels = [f"Period {i+1}" for i in range(PERIODS_PER_DAY)]
    num_cols      = 1 + PERIODS_PER_DAY   # Day col + period cols

    first_sheet_removed = False
    for cls in classes:
        ws = wb.create_sheet(title=sanitize(cls))
        if not first_sheet_removed:
            wb.remove(first_sheet)
            first_sheet_removed = True

        # Column widths
        ws.column_dimensions["A"].width = 10
        for col in range(2, num_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        # Title row
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=num_cols)
        title_cell            = ws.cell(row=1, column=1)
        title_cell.value      = f"Class {cls} — Timetable"
        title_cell.font       = Font(bold=True, size=13, color="FFFFFF")
        title_cell.fill       = PatternFill("solid", fgColor="1A3A5C")
        title_cell.alignment  = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        # Header row (Period labels)
        ws.cell(row=2, column=1).value = "Day"
        for i, lbl in enumerate(period_labels):
            ws.cell(row=2, column=i + 2).value = lbl
        _style_header_row(ws, 2, num_cols)
        ws.row_dimensions[2].height = 22

        # Data rows
        for d, day in enumerate(DAYS):
            row_num = d + 3
            ws.cell(row=row_num, column=1).value = day
            ws.cell(row=row_num, column=1).font  = Font(bold=True, size=10)
            ws.cell(row=row_num, column=1).fill  = PatternFill("solid", fgColor="DCE8F5")
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center",
                                                                  vertical="center")
            ws.cell(row=row_num, column=1).border = _make_border()
            ws.row_dimensions[row_num].height = 42

            for p in range(PERIODS_PER_DAY):
                col_num = p + 2
                c       = timetable[cls][d][p]
                if c:
                    subject = sanitize(c.get("subject", ""))
                    teacher = sanitize(c.get("teacher", ""))
                    ws.cell(row=row_num, column=col_num).value = f"{subject}\n{teacher}"
                else:
                    subject = ""
                    ws.cell(row=row_num, column=col_num).value = ""
                _style_data_cell(ws, row_num, col_num, subject, d)

        # Freeze header rows
        ws.freeze_panes = "B3"

    wb.save(xlsx_path)

    empty_report = build_empty_report(timetable, classes)

    summary = (
        f"Timetable generated successfully!\n"
        f"Empty slots : {best_empty}\n"
        f"Saved to    : {xlsx_path}\n"
        f"({len(classes)} class sheets)\n"
        f"backend_details.json updated"
    )

    if empty_report:
        summary += f"\n\n{empty_report}"

    return summary


# -------------------------------------------------------
# WORKER PROCESS ENTRY POINT
# -------------------------------------------------------

def _worker_process(msc_data, backend_data_existing, classes,
                    teacher_class_map, progress_queue, max_attempts):
    _run_attempts(msc_data, backend_data_existing, classes,
                  teacher_class_map, progress_queue, max_attempts)


# -------------------------------------------------------
# ASYNC ENTRY POINT
# -------------------------------------------------------

def generate_timetable_async(on_progress=None, on_done=None, on_error=None,
                             max_attempts=MAX_ATTEMPTS):

    def _watcher():
        try:
            if not os.path.exists(MSC_FILE):
                if on_error:
                    on_error("msc.json not found!")
                return

            with open(MSC_FILE) as f:
                msc_data = json.load(f)

            # ── Overbook check ────────────────────────────────────────────
            # Total slots in a week = DAYS × PERIODS_PER_DAY = 6 × 8 = 48.
            # If any teacher is assigned more than 48 periods total across
            # all classes, it is physically impossible to schedule them.
            WEEK_MAX = len(DAYS) * PERIODS_PER_DAY
            overbooked = []
            for teacher, info in msc_data.items():
                total = sum(info.get("classes", {}).values())
                if total > WEEK_MAX:
                    overbooked.append((teacher, total))

            if overbooked:
                lines = "\n".join(
                    f"  • {t}  —  {n} periods assigned  (max {WEEK_MAX})"
                    for t, n in overbooked
                )
                if on_error:
                    on_error(
                        f"Overbooked teacher schedule detected!\n\n"
                        f"The following teacher(s) have more periods assigned\n"
                        f"than there are slots in a week ({WEEK_MAX} total):\n\n"
                        f"{lines}\n\n"
                        f"Please go to Manage Teachers → Manage Teacher's Schedule\n"
                        f"and reduce their period counts before generating."
                    )
                return
            # ─────────────────────────────────────────────────────────────

            if os.path.exists(BACKEND_FILE):
                with open(BACKEND_FILE) as f:
                    backend_data_existing = json.load(f)
            else:
                backend_data_existing = {}

            classes = set()
            for info in msc_data.values():
                for cls in info.get("classes", {}).keys():
                    classes.add(cls)
            classes = sort_classes(list(classes))

            teacher_class_map = build_teacher_class_map(msc_data)

            progress_queue = multiprocessing.Queue()

            proc = multiprocessing.Process(
                target=_worker_process,
                args=(msc_data, backend_data_existing, classes,
                      teacher_class_map, progress_queue, max_attempts),
                daemon=True
            )
            proc.start()

            while True:
                msg = progress_queue.get()

                if "error" in msg:
                    proc.join()
                    if on_error:
                        on_error(msg["error"])
                    return

                if "done" in msg:
                    proc.join()
                    result = _build_excel(
                        msg["timetable"], msg["backend"],
                        msg["empty"], msc_data, classes
                    )
                    if on_done:
                        on_done(result)
                    return

                if on_progress:
                    on_progress(msg["attempt"], msg["total"], msg["empty"])

        except Exception as exc:
            if on_error:
                on_error(str(exc))

    t = threading.Thread(target=_watcher, daemon=True)
    t.start()
    return t