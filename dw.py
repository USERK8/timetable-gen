# dw.py

import os, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from paths import BACKEND_FILE

DAYS            = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
PERIODS_PER_DAY = 8
DOWNLOADS       = os.path.join(os.path.expanduser("~"), "Downloads")


# -------------------------------------------------------
# HELPERS
# -------------------------------------------------------

def sanitize(text):
    if not text:
        return ""
    text = str(text)
    result = []
    for ch in text:
        code = ord(ch)
        if 32 <= code <= 126:
            result.append(ch)
        elif code in (0x2014, 0x2013): result.append("-")
        elif code in (0x2018, 0x2019): result.append("'")
        elif code in (0x201C, 0x201D): result.append('"')
        elif code == 0x2026:           result.append("...")
        elif code == 0x00A0:           result.append(" ")
    return "".join(result).strip()


def _border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_border():
    t = Side(style="medium", color="888888")
    return Border(left=t, right=t, top=t, bottom=t)

def _header_fill():
    return PatternFill("solid", fgColor="1A3A5C")   # dark navy — teacher names col + period headers

def _day_title_fill():
    return PatternFill("solid", fgColor="0D2137")   # deeper navy for the day title row

def _busy_fill():
    return PatternFill("solid", fgColor="D4EDDA")   # soft green — has a class

def _free_fill(row_i):
    return PatternFill("solid", fgColor="F7F9FC" if row_i % 2 == 0 else "FFFFFF")

def _teacher_fill():
    return PatternFill("solid", fgColor="1E3A54")   # left column — teacher name


# -------------------------------------------------------
# LOAD BACKEND
# -------------------------------------------------------

def load_teacher_grids():
    if not os.path.exists(BACKEND_FILE):
        return None, (
            f"backend_details.json not found!\n"
            f"Expected: {BACKEND_FILE}\n"
            f"Generate the class-wise timetable first."
        )
    try:
        with open(BACKEND_FILE) as f:
            data = json.load(f)
    except Exception as e:
        return None, f"Error reading backend_details.json: {e}"

    grids = {}
    for teacher, info in data.items():
        if not teacher or teacher.strip() in {"—", "-", ""}:
            continue
        subject = sanitize(info.get("subject", ""))
        grid    = info.get("grid", [])

        normalized = []
        for day_idx in range(len(DAYS)):
            if day_idx < len(grid):
                row = list(grid[day_idx]) + [None] * PERIODS_PER_DAY
                normalized.append(row[:PERIODS_PER_DAY])
            else:
                normalized.append([None] * PERIODS_PER_DAY)

        grids[teacher] = {"subject": subject, "grid": normalized}

    return grids, None


# -------------------------------------------------------
# GENERATE DAY-WISE EXCEL
#
# Layout per sheet (one sheet = one day):
#
#   Row 1  : Day title spanning all columns
#   Row 2  : "Teacher  |  Subject  |  P1  |  P2  | ... | P8"
#   Row 3+ : teacher name | subject | class or blank per period
# -------------------------------------------------------

def generate_daywise_pdf():   # name kept so pet.py needs no change
    grids, err = load_teacher_grids()
    if err:
        return err
    if not grids:
        return "No teacher data found in backend_details.json."

    sorted_teachers = sorted(grids.keys())
    xlsx_path       = os.path.join(DOWNLOADS, "daywise_timetable.xlsx")

    wb          = Workbook()
    first_sheet = wb.active
    first_removed = False

    num_cols = 2 + PERIODS_PER_DAY   # Teacher | Subject | P1..P8

    for day_idx, day_name in enumerate(DAYS):
        ws = wb.create_sheet(title=day_name)
        if not first_removed:
            wb.remove(first_sheet)
            first_removed = True

        # ── Column widths ──────────────────────────────────────────
        ws.column_dimensions["A"].width = 24   # Teacher name
        ws.column_dimensions["B"].width = 16   # Subject
        for col in range(3, num_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12

        # ── Row 1: Day title ───────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=num_cols)
        tc           = ws.cell(row=1, column=1)
        tc.value     = f"{day_name} — Period Schedule"
        tc.font      = Font(bold=True, size=14, color="FFFFFF")
        tc.fill      = _day_title_fill()
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.border    = _border()
        ws.row_dimensions[1].height = 28

        # ── Row 2: Header ──────────────────────────────────────────
        headers = ["Teacher", "Subject"] + [f"Period {p+1}" for p in range(PERIODS_PER_DAY)]
        for col, hdr in enumerate(headers, start=1):
            cell           = ws.cell(row=2, column=col)
            cell.value     = hdr
            cell.font      = Font(bold=True, size=10, color="FFFFFF")
            cell.fill      = _header_fill()
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border    = _border()
        ws.row_dimensions[2].height = 22

        # ── Data rows: one per teacher ─────────────────────────────
        for row_i, teacher in enumerate(sorted_teachers):
            info    = grids[teacher]
            subject = info["subject"] or "—"
            day_row = info["grid"][day_idx]   # list of 8 class-or-None values
            xl_row  = row_i + 3

            ws.row_dimensions[xl_row].height = 30

            # Teacher name cell
            tc           = ws.cell(row=xl_row, column=1)
            tc.value     = sanitize(teacher)
            tc.font      = Font(bold=True, size=10, color="FFFFFF")
            tc.fill      = _teacher_fill()
            tc.alignment = Alignment(horizontal="left", vertical="center",
                                     indent=1)
            tc.border    = _border()

            # Subject cell
            sc           = ws.cell(row=xl_row, column=2)
            sc.value     = sanitize(subject)
            sc.font      = Font(size=9, color="AAAAAA", italic=True)
            sc.fill      = PatternFill("solid", fgColor="162A3E")
            sc.alignment = Alignment(horizontal="center", vertical="center")
            sc.border    = _border()

            # Period cells
            for p in range(PERIODS_PER_DAY):
                col  = p + 3
                cls  = day_row[p]
                cell = ws.cell(row=xl_row, column=col)

                if cls:
                    cell.value     = sanitize(cls)
                    cell.fill      = _busy_fill()
                    cell.font      = Font(size=10, bold=True, color="155724")
                else:
                    cell.value     = ""
                    cell.fill      = _free_fill(row_i)
                    cell.font      = Font(size=9, color="CCCCCC")

                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = _border()

        # Freeze teacher + subject columns and header row
        ws.freeze_panes = "C3"

    wb.save(xlsx_path)

    return (
        f"Day-wise Excel generated successfully!\n"
        f"{xlsx_path}\n"
        f"({len(DAYS)} day sheets, {len(sorted_teachers)} teachers)"
    )