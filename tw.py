# tw.py

import os, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from paths import BACKEND_FILE

DAYS            = ["Mon","Tue","Wed","Thu","Fri","Sat"]
PERIODS_PER_DAY = 8
DOWNLOADS       = os.path.join(os.path.expanduser("~"), "Downloads")


# -------------------------------------------------------
# SANITIZE
# -------------------------------------------------------

def sanitize(text):
    if not text:
        return ""
    text = str(text)
    result = []
    for ch in text:
        code = ord(ch)
        if 32 <= code <= 126:
            result.append(ch)
        elif code in (0x2014, 0x2013): result.append("-")
        elif code in (0x2018, 0x2019): result.append("'")
        elif code in (0x201C, 0x201D): result.append('"')
        elif code == 0x2026:           result.append("...")
        elif code == 0x00A0:           result.append(" ")
    return "".join(result).strip()


# -------------------------------------------------------
# EXCEL STYLE HELPERS
# -------------------------------------------------------

def _border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_fill():
    return PatternFill("solid", fgColor="2E4057")

def _busy_fill():
    return PatternFill("solid", fgColor="D4EDDA")   # soft green = has a class

def _free_fill(i):
    return PatternFill("solid", fgColor="F7F9FC" if i % 2 == 0 else "FFFFFF")


# -------------------------------------------------------
# LOAD TEACHER GRIDS
# -------------------------------------------------------

def load_teacher_grids():
    if not os.path.exists(BACKEND_FILE):
        return None, (
            f"backend_details.json not found!\n"
            f"Expected: {BACKEND_FILE}\n"
            f"Generate the timetable first."
        )
    try:
        with open(BACKEND_FILE) as f:
            data = json.load(f)
    except Exception as e:
        return None, f"Error reading backend_details.json: {e}"

    grids = {}
    for teacher, info in data.items():
        if not teacher or teacher.strip() in {"—", "-", ""}:
            continue
        subject = sanitize(info.get("subject", ""))
        grid    = info.get("grid", [])

        normalized = []
        for day_idx in range(len(DAYS)):
            if day_idx < len(grid):
                row = list(grid[day_idx]) + [None] * PERIODS_PER_DAY
                normalized.append(row[:PERIODS_PER_DAY])
            else:
                normalized.append([None] * PERIODS_PER_DAY)

        grids[teacher] = {"subject": subject, "grid": normalized}

    return grids, None


# -------------------------------------------------------
# GENERATE TEACHER-WISE EXCEL
# -------------------------------------------------------

def generate_teacherwise_pdf():   # kept same name so pet.py needs no change
    grids, err = load_teacher_grids()
    if err:
        return err
    if not grids:
        return "No teacher data found in backend_details.json."

    sorted_teachers = sorted(grids.keys())
    xlsx_path       = os.path.join(DOWNLOADS, "teacherwise_timetable.xlsx")
    wb              = Workbook()
    first_sheet     = wb.active
    first_removed   = False

    period_labels = [f"P{i+1}" for i in range(PERIODS_PER_DAY)]
    num_cols      = 1 + PERIODS_PER_DAY

    for teacher in sorted_teachers:
        info    = grids[teacher]
        subject = info["subject"] or "—"
        grid    = info["grid"]

        sheet_name = sanitize(teacher)[:31]
        ws = wb.create_sheet(title=sheet_name)
        if not first_removed:
            wb.remove(first_sheet)
            first_removed = True

        # Column widths
        ws.column_dimensions["A"].width = 10
        for col in range(2, num_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

        # Title row
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=num_cols)
        tc           = ws.cell(row=1, column=1)
        tc.value     = f"{teacher}  ({subject})"
        tc.font      = Font(bold=True, size=12, color="FFFFFF")
        tc.fill      = PatternFill("solid", fgColor="1A3A5C")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        # Header row
        ws.cell(row=2, column=1).value = "Day"
        for i, lbl in enumerate(period_labels):
            ws.cell(row=2, column=i + 2).value = lbl

        hdr_font  = Font(bold=True, color="FFFFFF", size=10)
        hdr_align = Alignment(horizontal="center", vertical="center")
        for col in range(1, num_cols + 1):
            cell            = ws.cell(row=2, column=col)
            cell.font       = hdr_font
            cell.fill       = _hdr_fill()
            cell.alignment  = hdr_align
            cell.border     = _border()
        ws.row_dimensions[2].height = 20

        # Data rows
        for d, day_name in enumerate(DAYS):
            row_num = d + 3
            ws.row_dimensions[row_num].height = 32

            day_cell           = ws.cell(row=row_num, column=1)
            day_cell.value     = day_name
            day_cell.font      = Font(bold=True, size=10)
            day_cell.fill      = PatternFill("solid", fgColor="DCE8F5")
            day_cell.alignment = Alignment(horizontal="center", vertical="center")
            day_cell.border    = _border()

            for p in range(PERIODS_PER_DAY):
                col_num = p + 2
                cls     = grid[d][p]
                cell    = ws.cell(row=row_num, column=col_num)

                if cls:
                    cell.value     = sanitize(cls)
                    cell.fill      = _busy_fill()
                    cell.font      = Font(size=10, color="155724", bold=True)
                else:
                    cell.value     = ""
                    cell.fill      = _free_fill(d)
                    cell.font      = Font(size=9, color="AAAAAA")

                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
                cell.border    = _border()

        ws.freeze_panes = "B3"

    wb.save(xlsx_path)

    return (
        f"Teacher-wise Excel generated successfully!\n"
        f"{xlsx_path}\n"
        f"({len(sorted_teachers)} teacher sheets)"
    )